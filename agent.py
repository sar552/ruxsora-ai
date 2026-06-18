#!/usr/bin/env python3
"""
ERPNext Telegram yordamchi-bot  (Claude API + tool use)

Foydalanuvchi Telegram'da oddiy savol yozadi, masalan:
  - "Kecha kim sales invoice kiritgan?"
  - "Hozir invoys qilinmagan delivery note'lar qancha?"
  - "Bu oy eng katta xaridlar kim qildi?"
Bot savolni Claude'ga uzatadi. Claude kerak bo'lsa ERPNext'dan
ma'lumot olib keladigan tool'ni o'zi chaqiradi, natijani tahlil qilib
o'zbekcha javob qaytaradi.

XAVFSIZLIK:
  * Bot ERPNext'ga FAQAT O'QISH (GET) so'rovlarini yuboradi. Hech narsa
    yaratmaydi/o'chirmaydi/o'zgartirmaydi.
  * Faqat ALLOWED_USERS ro'yxatidagi odamlar bot bilan gaplasha oladi.

Kerak:  pip install requests
"""

import io
import os
import re
import ast
import csv
import json
import time
import threading
import traceback
import requests
from pathlib import Path
from datetime import datetime

try:
    from zoneinfo import ZoneInfo
except ImportError:  # Python < 3.9
    ZoneInfo = None

# ===================== SOZLAMALAR (.env dan yuklanadi) =====================
# Barcha maxfiy qiymatlar agent.py yonidagi .env faylida turadi:
#   ERP_URL, ERP_KEY, ERP_SECRET, ANTHROPIC_KEY, MODEL, TG_TOKEN,
#   COMPANY (ixtiyoriy), ALLOWED_USERS (ro'yxat).
# .env qiymatlari Python sintaksisida ("matn" yoki [123, 456]) yozilgan va
# ast.literal_eval bilan XAVFSIZ o'qiladi (eval ISHLATILMAYDI).
#   * MODEL: eng arzoni "claude-haiku-4-5-20251001". Murakkabroq tahlil uchun
#     "claude-sonnet-4-6" qo'yish mumkin.
#   * ALLOWED_USERS: faqat shu Telegram ID'lar bot bilan gaplasha oladi
#     (o'z ID'ingizni bilish uchun @userinfobot ga yozing).
#   * COMPANY bo'sh bo'lsa, ERPNext'da bitta kompaniya bo'lsa avtomatik aniqlanadi.
# ===========================================================================


def _coerce(val):
    """'.env' qiymatini Python obyektiga aylantiradi (matn/son/ro'yxat).
    literal_eval bo'lmasa, oddiy tirnoqsiz matn sifatida qaytaradi."""
    try:
        return ast.literal_eval(val)
    except (ValueError, SyntaxError):
        return val.strip().strip("\"'")


def _load_env(path=None):
    """agent.py yonidagi .env ni o'qib, sozlamalarni shu modul globallariga yuklaydi.
    Operatsion tizim environment o'zgaruvchisi bo'lsa, u .env dan ustun turadi
    (serverda systemd orqali berishni osonlashtiradi)."""
    env_path = Path(path) if path else Path(__file__).with_name(".env")
    if not env_path.exists():
        raise SystemExit(
            f"DIQQAT: .env fayli topilmadi: {env_path}\n"
            "agent.py yonida .env bo'lishi kerak (ERP_URL, ERP_KEY, ... bilan)."
        )
    g = globals()
    for raw in env_path.read_text(encoding="utf-8").splitlines():
        line = raw.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        key = key.strip()
        g[key] = _coerce(os.environ[key]) if key in os.environ else _coerce(val.strip())


# Standart (bo'sh) qiymatlar — IDE/statik tahlil ko'rishi uchun. Haqiqiy
# qiymatlarni _load_env() quyida .env dan o'qib, shularning ustiga yozadi.
ERP_URL = ERP_KEY = ERP_SECRET = ANTHROPIC_KEY = MODEL = TG_TOKEN = COMPANY = ""
# Ixtiyoriy (yangi) sozlamalar:
#   MODEL_SMART  — tahlil/CFO savollari uchun kuchliroq model (masalan claude-sonnet-4-6).
#                  Bo'sh bo'lsa hamma narsa uchun MODEL ishlatiladi.
#   TIMEZONE     — "bugun/kecha/bu oy" qaysi vaqt mintaqasida hisoblanishi (default Asia/Tashkent).
#   DOCUMENTS_DIR— tayyor hisobot fayllari (balans/pnl PDF) turadigan papka.
MODEL_SMART = TIMEZONE = DOCUMENTS_DIR = ""
ALLOWED_USERS = []

_load_env()

# Majburiy sozlamalar to'g'ri yuklanganini tekshiramiz — bo'lmasa darhol to'xtaymiz.
_required = ["ERP_URL", "ERP_KEY", "ERP_SECRET", "ANTHROPIC_KEY",
             "MODEL", "TG_TOKEN", "ALLOWED_USERS"]
_missing = [k for k in _required if not globals().get(k)]
if _missing:
    raise SystemExit("DIQQAT: .env da quyidagilar yo'q yoki bo'sh: " + ", ".join(_missing))
if not isinstance(ALLOWED_USERS, (list, tuple, set)):
    raise SystemExit("DIQQAT: .env dagi ALLOWED_USERS ro'yxat bo'lishi kerak, masalan: [123, 456]")
COMPANY = globals().get("COMPANY") or ""        # bo'sh => avtomatik aniqlanadi
MODEL_SMART = globals().get("MODEL_SMART") or ""  # bo'sh => MODEL ishlatiladi
TIMEZONE = globals().get("TIMEZONE") or "Asia/Tashkent"
DOCUMENTS_DIR = globals().get("DOCUMENTS_DIR") or str(Path(__file__).with_name("documents"))
CURRENCY = globals().get("CURRENCY") or ""      # kompaniya hisob valyutasi (avtomatik aniqlanadi)

ERP_HEADERS = {"Authorization": f"token {ERP_KEY}:{ERP_SECRET}"}


def _now():
    """Sozlangan vaqt mintaqasidagi hozirgi vaqt. 'bugun/kecha' shunga asoslanadi."""
    if ZoneInfo is not None:
        try:
            return datetime.now(ZoneInfo(TIMEZONE))
        except Exception:
            pass
    return datetime.now()


def _log(*parts):
    """Stdout'ga vaqt belgili audit yozuvi (systemd journal yig'ib oladi)."""
    print(f"[{_now():%Y-%m-%d %H:%M:%S}]", *parts, flush=True)


def _autodetect_company():
    """COMPANY bo'sh bo'lsa, ERPNext'da bitta kompaniya bo'lsa uni avtomatik oladi.
    Shu bilan birga kompaniyaning hisob VALYUTASINI (CURRENCY) ham aniqlaydi."""
    global COMPANY, CURRENCY
    try:
        r = requests.get(f"{ERP_URL}/api/resource/Company",
                         headers=ERP_HEADERS,
                         params={"fields": json.dumps(["name", "default_currency"]),
                                 "limit_page_length": 20},
                         timeout=30)
        r.raise_for_status()
        comps = r.json().get("data", [])
        if not COMPANY:
            if len(comps) == 1:
                COMPANY = comps[0]["name"]
                print(f"Kompaniya avtomatik aniqlandi: {COMPANY}")
            elif len(comps) > 1:
                print(f"DIQQAT: {len(comps)} ta kompaniya bor. SOZLAMALAR'da COMPANY ni to'ldiring.")
        if not CURRENCY:
            match = next((c for c in comps if c.get("name") == COMPANY), None) or (comps[0] if comps else None)
            if match and match.get("default_currency"):
                CURRENCY = match["default_currency"]
                print(f"Hisob valyutasi aniqlandi: {CURRENCY}")
    except Exception as e:
        print("Kompaniya/valyutani aniqlab bo'lmadi:", e)


# ============ ERPNext O'QISH funksiyalari (tool'lar shu yerda bajariladi) =====
# Child table (qator) doctype'lari — bularni to'g'ridan-to'g'ri /resource orqali
# o'qib bo'lmaydi, frappe.client.get_list method orqali olamiz.
CHILD_DOCTYPES = {
    "Sales Invoice Item": "Sales Invoice",
    "Purchase Invoice Item": "Purchase Invoice",
    "Sales Order Item": "Sales Order",
    "Purchase Order Item": "Purchase Order",
    "Delivery Note Item": "Delivery Note",
    "Purchase Receipt Item": "Purchase Receipt",
}


def erp_list(doctype, filters=None, fields=None, limit=50, order_by=None):
    """Berilgan doctype bo'yicha hujjatlar ro'yxatini oladi (faqat o'qish).
    Child table bo'lsa frappe.client.get_list method orqali oladi."""
    if doctype in CHILD_DOCTYPES:
        # child table — method API orqali, parent maydonini ham qo'shamiz
        f = list(fields or ["name"])
        if "parent" not in f:
            f.append("parent")
        params = {
            "doctype": doctype,
            "filters": json.dumps(filters or []),
            "fields": json.dumps(f),
            "limit_page_length": limit,
            "parent": CHILD_DOCTYPES[doctype],
        }
        if order_by:
            params["order_by"] = order_by
        r = requests.get(f"{ERP_URL}/api/method/frappe.client.get_list",
                         headers=ERP_HEADERS, params=params, timeout=90)
        r.raise_for_status()
        return r.json().get("message", [])

    params = {
        "filters": json.dumps(filters or []),
        "fields": json.dumps(fields or ["name"]),
        "limit_page_length": limit,
    }
    if order_by:
        params["order_by"] = order_by
    r = requests.get(f"{ERP_URL}/api/resource/{doctype}",
                     headers=ERP_HEADERS, params=params, timeout=60)
    r.raise_for_status()
    return r.json().get("data", [])


def erp_get_doc(doctype, name):
    """Bitta hujjatning to'liq ma'lumotini oladi."""
    r = requests.get(f"{ERP_URL}/api/resource/{doctype}/{name}",
                     headers=ERP_HEADERS, timeout=60)
    r.raise_for_status()
    return r.json().get("data", {})


def erp_ledger(party=None, account=None, voucher_no=None, voucher_type=None,
               from_date=None, to_date=None, party_type=None, limit=50, filters=None):
    """
    GL Entry (bosh kitob) — TRANZAKSIYALARNING chuqur tafsiloti: pul qaysi hisobdan
    qaysi hisobga ketgani, kim bilan (party), debet/kredit, qaysi hujjat orqali.
      account     — hisob (qism nom ham bo'ladi, 'like' bilan qidiriladi), masalan 'Банк' yoki '1311'
      party       — mijoz/yetkazib beruvchi/xodim nomi
      voucher_no  — aniq hujjat raqami (masalan to'lov yoki invoys)
      voucher_type— Payment Entry / Sales Invoice / Journal Entry / Purchase Invoice ...
      from_date/to_date — sana oralig'i (posting_date)
    Natijada har posting: account, against (qarama-qarshi hisob), party, debit, credit,
    voucher_type, voucher_no, against_voucher (bog'liq invoys), remarks.
    """
    f = list(filters or [])
    f.append(["is_cancelled", "=", 0])
    if COMPANY:
        f.append(["company", "=", COMPANY])
    if party:
        f.append(["party", "=", party])
    if party_type:
        f.append(["party_type", "=", party_type])
    if account:
        f.append(["account", "like", f"%{account}%"])
    if voucher_no:
        f.append(["voucher_no", "=", voucher_no])
    if voucher_type:
        f.append(["voucher_type", "=", voucher_type])
    if from_date and to_date:
        f.append(["posting_date", "between", [from_date, to_date]])
    elif from_date:
        f.append(["posting_date", ">=", from_date])
    elif to_date:
        f.append(["posting_date", "<=", to_date])
    fields = ["posting_date", "account", "against", "party_type", "party",
              "debit", "credit", "voucher_type", "voucher_no",
              "against_voucher_type", "against_voucher", "remarks"]
    params = {
        "filters": json.dumps(f),
        "fields": json.dumps(fields),
        "limit_page_length": min(int(limit or 50), 200),
        "order_by": "posting_date desc, creation desc",
    }
    r = requests.get(f"{ERP_URL}/api/resource/GL Entry",
                     headers=ERP_HEADERS, params=params, timeout=90)
    r.raise_for_status()
    return r.json().get("data", [])


def erp_aggregate(doctype, group_by, metric="sum", field="base_grand_total",
                  filters=None, order="desc", limit=50):
    """
    Guruhlash + yig'ish (server tomonida hisoblanadi — kam token, aniq natija).
    Masalan: mijoz bo'yicha umumiy sotuv => doctype="Sales Invoice",
             group_by="customer", metric="sum", field="base_grand_total".
    metric: sum | count | avg | max | min
    """
    if doctype in CHILD_DOCTYPES:
        # Child table'larni /resource group_by bilan yig'ib bo'lmaydi — tushunarli xato qaytaramiz.
        return {"error": f"'{doctype}' — bu child (qator) table. Uni erp_aggregate bilan "
                         f"guruhlab bo'lmaydi. Tovar darajasidagi tahlil uchun erp_profit_breakdown ishlat."}
    expr = "count(name)" if metric == "count" else f"{metric}(`{field}`)"
    params = {
        "fields": json.dumps([group_by, f"{expr} as value"]),
        "filters": json.dumps(filters or []),
        "group_by": group_by,
        "order_by": f"{expr} {order}",
        "limit_page_length": limit,
    }
    r = requests.get(f"{ERP_URL}/api/resource/{doctype}",
                     headers=ERP_HEADERS, params=params, timeout=90)
    r.raise_for_status()
    return r.json().get("data", [])


def _run_query_report(report_name, filters=None):
    """Past darajali: query report'ni ishga tushirib, xom 'message' ni qaytaradi.
    ignore_prepared_report=1 — 'Prepared Report' qilib sozlangan hisobotlarni
    (masalan Balance Sheet) fonda kutmasdan, DARHOL sinxron hisoblashga majbur qiladi."""
    r = requests.get(f"{ERP_URL}/api/method/frappe.desk.query_report.run",
                     headers=ERP_HEADERS,
                     params={"report_name": report_name,
                             "filters": json.dumps(filters or {}),
                             "ignore_prepared_report": 1},
                     timeout=180)
    r.raise_for_status()
    return r.json().get("message", {})


def _rows_as_dicts(message):
    """Hisobot qatorlarini (dict yoki list bo'lishidan qat'i nazar) dict ro'yxatiga aylantiradi."""
    fieldnames = []
    for c in message.get("columns", []):
        if isinstance(c, dict):
            fieldnames.append(c.get("fieldname") or c.get("label"))
        else:
            fieldnames.append(str(c).split(":")[0].strip())
    out = []
    for row in message.get("result", []):
        if isinstance(row, dict):
            out.append(row)
        elif isinstance(row, (list, tuple)):
            out.append({fieldnames[i]: row[i] for i in range(min(len(fieldnames), len(row)))})
    return out


def erp_run_report(report_name, filters=None):
    """ERPNext tayyor hisobotini ishga tushiradi (masalan 'Gross Profit' — real foyda)."""
    msg = _run_query_report(report_name, filters)
    cols = []
    for c in msg.get("columns", []):
        cols.append(c.get("label") or c.get("fieldname") if isinstance(c, dict) else str(c).split(":")[0])
    return {"columns": cols, "rows": msg.get("result", [])[:400]}


def erp_profit_breakdown(from_date, to_date, by="item_code", customer=None, top=20,
                         company=None, rank_by="gross_profit", order="desc"):
    """
    Tovar/mijoz kesimida SOTUV va FOYDANI hisoblaydi (har element uchun: foyda, tushum, miqdor).
    'Gross Profit' hisobotini Invoice darajasida olib, so'ng Python ichida qayta yig'adi.
    Shu sababli "top sotilgan mahsulot", "eng kam sotilgan", "qaysi tovar ko'p foyda",
    "falon mijozning qaysi tovari" kabi savollarning hammasiga javob bera oladi.
      by      = "item_code" | "customer" | "customer_item"
      customer= faqat shu mijoz bilan cheklash (ixtiyoriy)
      rank_by = nima bo'yicha saralash: "gross_profit" (foyda) | "revenue" (tushum/savdo) | "qty" (miqdor)
                * "Top 10 mahsulot / ko'p sotilgan"  => rank_by="revenue" (yoki "qty")
                * "Eng ko'p foyda keltiradi"          => rank_by="gross_profit"
      order   = "desc" (ko'pdan kamga, top) | "asc" (kamdan ko'pga — "eng kam sotilgan" uchun)
    """
    comp = company or COMPANY
    if not comp:
        return {"error": "Bu hisobot uchun 'company' kerak. SOZLAMALAR'dagi COMPANY ni to'ldiring "
                         "yoki so'rovda kompaniya nomini ayting."}
    msg = _run_query_report("Gross Profit", {
        "company": comp, "from_date": from_date, "to_date": to_date, "group_by": "Invoice",
    })
    rows = _rows_as_dicts(msg)

    cur_customer = None       # invoice sarlavhasidagi mijozni keyingi tovar qatorlariga tarqatamiz
    agg = {}
    for r in rows:
        if r.get("customer"):
            cur_customer = r.get("customer")
        item = r.get("item_code")
        if not item:
            continue                      # subtotal/sarlavha qatorini o'tkazib yuboramiz
        # gross_profit bo'sh bo'lsa ham (tannarx kiritilmagan) tushum/miqdor bo'yicha tahlilga olamiz
        gp = r.get("gross_profit")
        cust = r.get("customer") or cur_customer
        if customer and cust != customer:
            continue
        if by == "customer":
            key = cust
        elif by == "customer_item":
            key = f"{cust} | {item}"
        else:
            key = item
        a = agg.setdefault(key, {"gross_profit": 0.0, "revenue": 0.0, "qty": 0.0})
        a["gross_profit"] += float(gp or 0)
        # Gross Profit hisobotida tushum ustuni 'selling_amount' (base_amount emas!)
        a["revenue"] += float(r.get("selling_amount") or r.get("base_amount") or 0)
        a["qty"] += float(r.get("qty") or 0)

    rank_key = rank_by if rank_by in ("gross_profit", "revenue", "qty") else "gross_profit"
    ranked = sorted(agg.items(), key=lambda kv: kv[1][rank_key],
                    reverse=(order != "asc"))[:top]
    if not ranked:
        return {"warning": "Ma'lumot topilmadi — bu davrda sotuv yo'q yoki hisobot bo'sh. "
                           "(Foyda bo'sh chiqsa, tovar tannarxi/valuation kiritilmagan bo'lishi mumkin.)",
                "items": []}
    return {"by": by, "customer": customer, "ranked_by": rank_key, "order": order,
            "note": "revenue=tushum/savdo, gross_profit=yalpi foyda, qty=miqdor. "
                    "Eslatma: bu yerda faqat shu davrda SOTILGAN tovarlar bor "
                    "(umuman sotilmagan/dead-stock tovarlar uchun Stock hisobotidan foydalaning).",
            "items": [dict(key=k, **v) for k, v in ranked]}


# ============ HISOBOTNI FAYL (Excel/CSV/PDF) qilib yuborish ===================
def _report_table(report_name, filters=None):
    """Hisobotni ishga tushirib, (sarlavhalar, qatorlar matritsasi) ko'rinishida qaytaradi.
    Moliyaviy hisobotlardagi 'indent' (ichki bandlar) bo'lsa, hisob nomiga bo'sh joy qo'shadi."""
    msg = _run_query_report(report_name, filters)
    columns = msg.get("columns", [])
    headers, fieldnames = [], []
    for c in columns:
        if isinstance(c, dict):
            fn = c.get("fieldname") or c.get("label") or ""
            headers.append(c.get("label") or fn)
            fieldnames.append(fn)
        else:
            label = str(c).split(":")[0].strip()
            headers.append(label)
            fieldnames.append(label)

    # qaysi ustun "nomi" ustuni (indentni shunga qo'llaymiz)
    name_fields = ("account", "account_name", "party", "item", "label")
    matrix = []
    for row in msg.get("result", []):
        if isinstance(row, dict):
            indent = int(row.get("indent") or 0)
            line = []
            for fn in fieldnames:
                v = row.get(fn, "")
                if indent and fn in name_fields and isinstance(v, str) and v:
                    v = ("    " * indent) + v
                line.append(v)
            matrix.append(line)
        elif isinstance(row, (list, tuple)):
            matrix.append([row[i] if i < len(row) else "" for i in range(len(fieldnames))])
        # None (ajratuvchi) qatorlarni tashlab yuboramiz
    return headers, matrix


def _build_xlsx(headers, matrix, sheet_title="Hisobot"):
    """openpyxl bilan .xlsx bayt-massivini yasaydi."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_title or "Hisobot")[:31]
    ws.append([str(h) for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in matrix:
        ws.append([_clean_cell(v) for v in row])
    # ustun kengligini sarlavha bo'yicha taxminan moslaymiz
    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(40, len(str(h)) + 4))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _clean_cell(v):
    """Excel katakchasi uchun qiymatni tozalaydi (raqam — raqam, qolgani matn)."""
    if v is None:
        return ""
    if isinstance(v, (int, float, str)):
        return v
    return str(v)


def _build_csv(headers, matrix):
    """CSV bayt-massivi (Excel to'g'ri ochishi uchun UTF-8 BOM bilan)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([str(h) for h in headers])
    for row in matrix:
        w.writerow(["" if v is None else v for v in row])
    return buf.getvalue().encode("utf-8-sig")


def _report_html(title, headers, matrix):
    """Hisobotdan sodda, chiroyli HTML jadval yasaydi (PDF uchun)."""
    def esc(x):
        return (str(x) if x is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    th = "".join(f"<th>{esc(h)}</th>" for h in headers)
    trs = []
    for row in matrix:
        tds = "".join(f"<td>{esc(v)}</td>" for v in row)
        trs.append(f"<tr>{tds}</tr>")
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
        body{{font-family:Arial,sans-serif;font-size:11px;color:#222}}
        h2{{margin:0 0 4px}} .sub{{color:#666;font-size:10px;margin-bottom:10px}}
        table{{border-collapse:collapse;width:100%}}
        th,td{{border:1px solid #ccc;padding:4px 6px;text-align:left}}
        th{{background:#f2f2f2}} td:not(:first-child){{text-align:right}}
        </style></head><body>
        <h2>{esc(title)}</h2>
        <div class="sub">{esc(COMPANY)} · {esc(title)}</div>
        <table><thead><tr>{th}</tr></thead><tbody>{''.join(trs)}</tbody></table>
        </body></html>"""


def _build_pdf(title, headers, matrix, orientation="Landscape"):
    """ERPNext'ning o'z PDF generatori (wkhtmltopdf) orqali PDF yasaydi.
    Mahalliy PDF kutubxonasi shart emas. Muvaffaqiyatsiz bo'lsa exception tashlaydi."""
    html = _report_html(title, headers, matrix)
    r = requests.post(f"{ERP_URL}/api/method/frappe.utils.print_format.report_to_pdf",
                      headers=ERP_HEADERS,
                      data={"html": html, "orientation": orientation},
                      timeout=120)
    r.raise_for_status()
    ctype = r.headers.get("content-type", "")
    if "pdf" not in ctype and not r.content[:4] == b"%PDF":
        raise RuntimeError(f"PDF o'rniga boshqa javob keldi ({ctype}): {r.text[:200]}")
    return r.content


# Moliyaviy hisobotlar uchun MAJBURIY (sukutdagi) filtrlar.
# Foydalanuvchi faqat sana oralig'ini beradi; qolganini bot avtomatik to'ldiradi.
#   Balance Sheet            -> Date Range, Monthly, accumulated_values YOQIQ
#   Profit and Loss Statement-> Date Range, Monthly, accumulated_values O'CHIQ
#   Cash Flow                -> Date Range, Monthly
FIN_REPORT_DEFAULTS = {
    "Balance Sheet": {"filter_based_on": "Date Range", "periodicity": "Monthly", "accumulated_values": 1},
    "Profit and Loss Statement": {"filter_based_on": "Date Range", "periodicity": "Monthly", "accumulated_values": 0},
    "Cash Flow": {"filter_based_on": "Date Range", "periodicity": "Monthly"},
}

# Foydalanuvchi/Claude turlicha atashi mumkin — kanonik ERPNext nomiga keltiramiz.
REPORT_ALIASES = {
    "balance sheet": "Balance Sheet",
    "balans": "Balance Sheet",
    "balanssheet": "Balance Sheet",
    "profit and loss": "Profit and Loss Statement",
    "profit and loss statement": "Profit and Loss Statement",
    "p&l": "Profit and Loss Statement",
    "pnl": "Profit and Loss Statement",
    "foyda va zarar": "Profit and Loss Statement",
    "foyda zarar": "Profit and Loss Statement",
    "cash flow": "Cash Flow",
    "cash flow statement": "Cash Flow",
    "pul oqimi": "Cash Flow",
}


def _canonical_report_name(name):
    # chiziqcha/ortiqcha bo'shliqlarni bo'sh joyga keltirib, alias'ni bardoshli izlaymiz
    key = " ".join((name or "").replace("-", " ").split()).lower()
    return REPORT_ALIASES.get(key, name)


def _apply_fin_defaults(report_name, filters):
    """Moliyaviy hisobot uchun majburiy filtrlarni qo'llaydi. Sana (period_start_date/
    period_end_date) foydalanuvchidan keladi, qolgan texnik filtrlar avtomatik."""
    f = dict(filters or {})
    f.setdefault("company", COMPANY)
    defaults = FIN_REPORT_DEFAULTS.get(report_name)
    if defaults:
        f.update(defaults)   # bu sozlamalar MAJBURIY (P&L da accumulated o'chiq bo'lib qoladi, h.k.)
    return f


def _report_data_text(report_name, headers, matrix, period=None,
                      limit_rows=300, limit_chars=45000):
    """Hisobot jadvalini Claude tahlil qila oladigan ixcham JSON-matnga aylantiradi."""
    payload = {
        "report": report_name,
        "period": period,
        "columns": [str(h) for h in headers],
        "rows": [[_clean_cell(v) for v in row] for row in matrix[:limit_rows]],
    }
    return json.dumps(payload, ensure_ascii=False, default=str)[:limit_chars]


def get_report_data(args):
    """Hisobot ma'lumotini FAYL yubormasdan matn ko'rinishida qaytaradi (tahlil uchun)."""
    report_name = _canonical_report_name(args["report_name"])
    filters = _apply_fin_defaults(report_name, args.get("filters"))
    if report_name in FIN_REPORT_DEFAULTS and not (filters.get("period_start_date") and filters.get("period_end_date")):
        return ("Bu hisobot uchun sana oralig'i kerak. period_start_date va "
                "period_end_date (YYYY-MM-DD) ber.")
    headers, matrix = _report_table(report_name, filters)
    if not matrix:
        return "Hisobot bo'sh — ma'lumot topilmadi. Filtrlar/sanani tekshiring."
    period = f"{filters.get('period_start_date')} .. {filters.get('period_end_date')}"
    return _report_data_text(report_name, headers, matrix, period)


def send_report_file(args, chat_id):
    """ERPNext hisobotini fayl qilib (xlsx/csv/pdf) Telegram'ga yuboradi.
    Claude shu tool orqali foydalanuvchiga haqiqiy fayl jo'nata oladi."""
    if chat_id is None:
        return "Fayl yuborib bo'lmadi: chat aniqlanmadi."
    report_name = _canonical_report_name(args["report_name"])
    filters = _apply_fin_defaults(report_name, args.get("filters"))
    if report_name in FIN_REPORT_DEFAULTS and not (filters.get("period_start_date") and filters.get("period_end_date")):
        return ("Bu hisobot uchun sana oralig'i kerak. Foydalanuvchidan qaysi davr ekanini "
                "aniqlab, filters ichida period_start_date va period_end_date (YYYY-MM-DD) ber.")
    fmt = (args.get("file_format") or "xlsx").lower()
    title = args.get("title") or report_name
    base = (args.get("filename") or report_name).replace(" ", "_").replace("/", "-")

    headers, matrix = _report_table(report_name, filters)
    if not matrix:
        return ("Hisobot bo'sh — yuboradigan ma'lumot topilmadi. "
                "Filtrlarni (sana/kompaniya) tekshiring.")

    caption = args.get("caption") or title
    try:
        if fmt == "csv":
            content, fname = _build_csv(headers, matrix), f"{base}.csv"
        elif fmt == "pdf":
            content, fname = _build_pdf(title, headers, matrix), f"{base}.pdf"
        else:
            content, fname = _build_xlsx(headers, matrix, report_name), f"{base}.xlsx"
    except Exception as e:
        # PDF/Excel yasashda muammo bo'lsa — CSV ga qaytamiz, baribir fayl ketsin
        content, fname = _build_csv(headers, matrix), f"{base}.csv"
        tg_send_document(chat_id, fname, content, caption=caption)
        return (f"'{fmt}' formatini yasab bo'lmadi ({e}). Buning o'rniga CSV yubordim: "
                f"{fname} ({len(matrix)} qator).")

    tg_send_document(chat_id, fname, content, caption=caption)
    period = f"{filters.get('period_start_date')} .. {filters.get('period_end_date')}"
    data_text = _report_data_text(report_name, headers, matrix, period)
    return (f"Fayl yuborildi: {fname} ({len(matrix)} qator, {len(headers)} ustun).\n"
            f"Shu hisobot ma'lumoti (kerak bo'lsa tahlil uchun):\n{data_text}")


# ============ MOLIYAVIY HISOBOT (Balans / P&L) — GL Entry datasidan =============
# Kompaniyaning OYLIK Balans va Foyda-Zarar hisobotini REAL ERPNext ma'lumotidan
# (GL Entry) quradi va kompaniyaning rus tilidagi hisob nomlari bilan, oylik
# ustunlarda, shablonga o'xshash ko'rinishda PDF/Excel qilib beradi.
# DIQQAT: standart "Balance Sheet"/"Profit and Loss" hisoboti moliyaviy yil
# sozlamasiga bog'liq (bu kompaniyada yillar ustma-ust => xato). Shuning uchun
# biz GL Entry'dan to'g'ridan-to'g'ri hisoblaymiz — moliyaviy yildan mustaqil.

_RU_MONTHS = ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
              "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
_ACCOUNTS_CACHE = {}   # company -> (accs, idx, children)


def _account_index(company):
    """Kompaniyaning butun hisob rejasini bir marta olib, keshlaydi."""
    if company in _ACCOUNTS_CACHE:
        return _ACCOUNTS_CACHE[company]
    accs = erp_list("Account",
                    filters=[["company", "=", company]],
                    fields=["name", "account_name", "root_type", "parent_account",
                            "is_group", "account_number", "lft"],
                    limit=2000, order_by="lft asc")
    idx = {a["name"]: a for a in accs}
    children = {}
    for a in accs:
        children.setdefault(a.get("parent_account"), []).append(a["name"])
    _ACCOUNTS_CACHE[company] = (accs, idx, children)
    return accs, idx, children


def _gl_balances(company, d1, d2):
    """[d1..d2] oralig'ida har hisob bo'yicha (debet − kredit) yig'indisi. Faqat o'qish."""
    params = {
        "fields": json.dumps(["account", "sum(debit) as dr", "sum(credit) as cr"]),
        "filters": json.dumps([["company", "=", company],
                               ["posting_date", "between", [d1, d2]],
                               ["is_cancelled", "=", 0]]),
        "group_by": "account",
        "limit_page_length": 0,
    }
    r = requests.get(f"{ERP_URL}/api/resource/GL Entry",
                     headers=ERP_HEADERS, params=params, timeout=120)
    r.raise_for_status()
    return {x["account"]: (float(x.get("dr") or 0) - float(x.get("cr") or 0))
            for x in r.json().get("data", [])}


def _iter_months(start, end):
    """'YYYY-MM-DD' oralig'ini oylik bo'laklarga ajratadi (ko'pi bilan 24 oy)."""
    import calendar
    y, m = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]), int(end[5:7])
    out = []
    while (y < ey) or (y == ey and m <= em):
        last = calendar.monthrange(y, m)[1]
        out.append({"label": f"{_RU_MONTHS[m - 1]} {y}",
                    "start": f"{y:04d}-{m:02d}-01", "end": f"{y:04d}-{m:02d}-{last:02d}"})
        m += 1
        if m > 12:
            m, y = 1, y + 1
        if len(out) > 24:
            break
    return out


def _sign(root_type):
    # Income/Liability/Equity kredit-normal (musbat ko'rsatish uchun belgini teskari)
    return -1.0 if root_type in ("Income", "Liability", "Equity") else 1.0


def build_financial_statement(kind, start, end, company=None):
    """kind='pnl' yoki 'balance'. Oylik ustunli, ierarxik hisobot tuzilmasini qaytaradi."""
    company = company or COMPANY
    if not company:
        return {"error": "COMPANY aniqlanmagan."}
    accs, idx, children = _account_index(company)
    if not accs:
        return {"error": f"'{company}' uchun hisob rejasi topilmadi."}
    months = _iter_months(start, end)
    roots = ("Income", "Expense") if kind == "pnl" else ("Asset", "Liability", "Equity")

    # Har oy uchun hisob qoldiqlari: P&L oy ichidagi aylanma, Balans — boshidan to oy oxirigacha (jami)
    monthly = []
    for mo in months:
        if kind == "pnl":
            monthly.append(_gl_balances(company, mo["start"], mo["end"]))
        else:
            monthly.append(_gl_balances(company, "1900-01-01", mo["end"]))

    def acc_value(name, i):
        a = idx[name]
        if a.get("is_group"):
            return sum(acc_value(c, i) for c in children.get(name, []))
        return monthly[i].get(name, 0.0) * _sign(a.get("root_type"))

    def depth(name):
        d, p = 0, idx[name].get("parent_account")
        while p and p in idx and d < 12:
            d += 1
            p = idx[p].get("parent_account")
        return d

    n = len(months)

    def nonzero(name):
        return any(abs(acc_value(name, i)) > 0.005 for i in range(n))

    rows = []

    def emit_tree(name):
        a = idx[name]
        if not nonzero(name):
            return
        vals = [acc_value(name, i) for i in range(n)]
        label = a.get("account_name") or a["name"]
        rows.append({"label": label, "indent": depth(name),
                     "kind": "group" if a.get("is_group") else "leaf", "values": vals})
        if a.get("is_group"):
            for c in children.get(name, []):
                emit_tree(c)

    def root_total(rt):
        return [sum(acc_value(nm, i) for nm in idx if idx[nm].get("root_type") == rt
                    and not idx[nm].get("parent_account"))
                for i in range(n)]
    # eng yuqori (parent yo'q) hisoblardan boshlab daraxtni chizamiz
    top_level = [a["name"] for a in accs if not a.get("parent_account")]

    if kind == "pnl":
        for nm in top_level:
            if idx[nm].get("root_type") in roots:
                emit_tree(nm)
        income_total = root_total("Income")
        expense_total = root_total("Expense")
        net = [income_total[i] - expense_total[i] for i in range(n)]
        rows.append({"label": "Итого доходы (Выручка)", "indent": 0, "kind": "total", "values": income_total})
        rows.append({"label": "Итого расходы", "indent": 0, "kind": "total", "values": expense_total})
        rows.append({"label": "Чистая прибыль", "indent": 0, "kind": "grand", "values": net})
        margin = [(net[i] / income_total[i] * 100 if income_total[i] else 0.0) for i in range(n)]
        rows.append({"label": "Рентабельность по чистой прибыли, %", "indent": 0,
                     "kind": "pct", "values": margin})
        title = "Отчет о прибылях и убытках (P&L)"
    else:
        for nm in top_level:
            if idx[nm].get("root_type") in roots:
                emit_tree(nm)
        asset_total = root_total("Asset")
        liab_total = root_total("Liability")
        equity_total = root_total("Equity")
        # Joriy davr (yopilmagan) foydasi kapitalga kiradi — aks holda balans balanslanmaydi.
        # Balans rejimida qoldiqlar JAMI (inception..oy oxiri), shuning uchun foyda ham jami.
        retained = [root_total("Income")[i] - root_total("Expense")[i] for i in range(n)]
        equity_with_pl = [equity_total[i] + retained[i] for i in range(n)]
        passiv = [liab_total[i] + equity_with_pl[i] for i in range(n)]
        rows.append({"label": "Прибыль/убыток текущего периода (нераспределённая)",
                     "indent": 1, "kind": "leaf", "values": retained})
        rows.append({"label": "Капитал (с учётом прибыли)", "indent": 0, "kind": "total",
                     "values": equity_with_pl})
        rows.append({"label": "ИТОГО Активы", "indent": 0, "kind": "grand", "values": asset_total})
        rows.append({"label": "ИТОГО Пассивы (Обязательства + Капитал)", "indent": 0,
                     "kind": "grand", "values": passiv})
        diff = [asset_total[i] - passiv[i] for i in range(n)]
        rows.append({"label": "Разница (Актив − Пассив)", "indent": 0, "kind": "total", "values": diff})
        title = "Бухгалтерский баланс"

    return {"kind": kind, "title": title, "company": company,
            "columns": [mo["label"] for mo in months], "rows": rows}


def _fmt_num(v):
    """1 234 567,89 ko'rinishida (rus formati) — manfiy bo'lsa minus bilan."""
    try:
        s = f"{float(v):,.2f}"            # 1,234,567.89
    except (TypeError, ValueError):
        return "" if v is None else str(v)
    s = s.replace(",", " ").replace(".", ",")  # -> 1 234 567,89
    return s


def _statement_matrix(stmt):
    """xlsx/csv uchun (sarlavhalar, matritsa)."""
    headers = ["Статья"] + stmt["columns"]
    matrix = []
    for r in stmt["rows"]:
        label = ("    " * r["indent"]) + r["label"]
        if r["kind"] == "pct":
            vals = [f"{v:.0f}%" for v in r["values"]]
        else:
            vals = r["values"]
        matrix.append([label] + list(vals))
    return headers, matrix


def _statement_html(stmt):
    """Shablonga o'xshash, rangli bo'limli HTML (PDF uchun)."""
    def esc(x):
        return (str(x) if x is not None else "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    th = "".join(f"<th>{esc(c)}</th>" for c in stmt["columns"])
    trs = []
    for r in stmt["rows"]:
        cls = r["kind"]
        pad = 6 + r["indent"] * 16
        if r["kind"] == "pct":
            cells = "".join(f"<td>{v:.0f}%</td>" for v in r["values"])
        else:
            cells = "".join(f"<td>{esc(_fmt_num(v))}</td>" for v in r["values"])
        trs.append(f'<tr class="{cls}"><td class="lbl" style="padding-left:{pad}px">'
                   f'{esc(r["label"])}</td>{cells}</tr>')
    return f"""<!doctype html><html><head><meta charset="utf-8"><style>
        body{{font-family:Arial,sans-serif;font-size:11px;color:#222}}
        h2{{margin:0 0 2px}} .sub{{color:#666;font-size:10px;margin-bottom:10px}}
        table{{border-collapse:collapse;width:100%}}
        th,td{{border:1px solid #ccc;padding:4px 6px;text-align:right;white-space:nowrap}}
        th{{background:#404040;color:#fff;text-align:right}} th:first-child{{text-align:left}}
        td.lbl{{text-align:left}}
        tr.group>td{{background:#f2f2f2;font-weight:bold}}
        tr.total>td{{font-weight:bold;border-top:2px solid #888}}
        tr.grand>td{{background:#fde9d9;font-weight:bold;border-top:2px solid #c0504d}}
        tr.pct>td{{font-style:italic;color:#555}}
        </style></head><body>
        <h2>{esc(stmt['title'])}</h2>
        <div class="sub">{esc(stmt['company'])} · {esc(stmt['columns'][0])} – {esc(stmt['columns'][-1])} · валюта: {esc(CURRENCY or 'USD')}</div>
        <table><thead><tr><th>Статья</th>{th}</tr></thead><tbody>{''.join(trs)}</tbody></table>
        </body></html>"""


def _statement_pdf(stmt):
    """HTML'ni ERPNext PDF generatori (wkhtmltopdf) orqali PDF qiladi."""
    html = _statement_html(stmt)
    r = requests.post(f"{ERP_URL}/api/method/frappe.utils.print_format.report_to_pdf",
                      headers=ERP_HEADERS, data={"html": html, "orientation": "Landscape"},
                      timeout=120)
    r.raise_for_status()
    if "pdf" not in r.headers.get("content-type", "") and r.content[:4] != b"%PDF":
        raise RuntimeError(f"PDF o'rniga boshqa javob: {r.text[:200]}")
    return r.content


def _statement_text(stmt, limit=40000):
    """Tahlil uchun ixcham JSON-matn (Claude o'qiydi)."""
    payload = {"title": stmt["title"], "currency": CURRENCY or "USD",
               "columns": stmt["columns"],
               "rows": [{"label": r["label"], "indent": r["indent"],
                         "values": [round(v, 2) if isinstance(v, (int, float)) else v
                                    for v in r["values"]]} for r in stmt["rows"]]}
    return json.dumps(payload, ensure_ascii=False, default=str)[:limit]


# "balans"/"pnl" so'zlarini kind'ga keltirish
_FIN_KIND = {
    "balance": "balance", "balans": "balance", "баланс": "balance", "balance sheet": "balance",
    "pnl": "pnl", "p&l": "pnl", "pl": "pnl", "foyda zarar": "pnl", "foyda-zarar": "pnl",
    "foyda va zarar": "pnl", "profit and loss": "pnl", "прибыль": "pnl",
}


def send_financial_statement(args, chat_id):
    """Balans yoki P&L hisobotini REAL ERPNext (GL) datasidan tuzib, oylik ustunli
    PDF/Excel qilib Telegram'ga yuboradi va tahlil uchun raqamlarni qaytaradi."""
    if chat_id is None:
        return "Fayl yuborib bo'lmadi: chat aniqlanmadi."
    kind = _FIN_KIND.get((args.get("document") or args.get("kind") or "").strip().lower())
    if kind is None:
        return "document 'balance' yoki 'pnl' bo'lishi kerak."
    start = args.get("period_start_date")
    end = args.get("period_end_date")
    if not (start and end):
        return ("Bu hisobot uchun sana oralig'i kerak. Foydalanuvchidan davrni aniqlab, "
                "period_start_date va period_end_date (YYYY-MM-DD) ber. Masalan oxirgi 4 oy.")
    try:
        stmt = build_financial_statement(kind, start, end)
    except Exception as e:
        return f"Hisobotni tuzishda xato: {e}"
    if stmt.get("error"):
        return stmt["error"]
    if not stmt.get("rows"):
        return "Bu davrda ma'lumot topilmadi (GL bo'sh). Sanani tekshiring."

    fmt = (args.get("file_format") or "pdf").lower()
    base = f"{kind}_{start}_{end}".replace("-", "")
    caption = args.get("caption") or f"{stmt['title']} · {stmt['columns'][0]}–{stmt['columns'][-1]}"
    headers, matrix = _statement_matrix(stmt)
    try:
        if fmt == "csv":
            content, fname = _build_csv(headers, matrix), f"{base}.csv"
        elif fmt in ("xlsx", "excel"):
            content, fname = _build_xlsx(headers, matrix, stmt["title"][:31]), f"{base}.xlsx"
        else:
            content, fname = _statement_pdf(stmt), f"{base}.pdf"
    except Exception as e:
        content, fname = _build_csv(headers, matrix), f"{base}.csv"
        tg_send_document(chat_id, fname, content, caption=caption)
        return (f"'{fmt}' yasab bo'lmadi ({e}). CSV yubordim: {fname}.\n"
                f"Ma'lumot:\n{_statement_text(stmt)}")
    tg_send_document(chat_id, fname, content, caption=caption)
    return (f"Fayl yuborildi: {fname} ({len(stmt['columns'])} oy, {len(stmt['rows'])} qator). "
            f"Real ERPNext (GL) ma'lumoti.\nTahlil uchun raqamlar:\n{_statement_text(stmt)}")


def run_tool(name, args, chat_id=None):
    """Claude chaqirgan tool'ni bajarib, natijani matn ko'rinishida qaytaradi."""
    try:
        if name == "send_financial_statement":
            return send_financial_statement(args, chat_id)
        if name == "send_report_file":
            return send_report_file(args, chat_id)
        if name == "get_report_data":
            return get_report_data(args)
        if name == "erp_list":
            data = erp_list(
                args["doctype"],
                args.get("filters"),
                args.get("fields"),
                args.get("limit", 50),
                args.get("order_by"),
            )
            return json.dumps(data, ensure_ascii=False, default=str)
        elif name == "erp_get_doc":
            data = erp_get_doc(args["doctype"], args["name"])
            return json.dumps(data, ensure_ascii=False, default=str)
        elif name == "erp_ledger":
            data = erp_ledger(
                args.get("party"), args.get("account"), args.get("voucher_no"),
                args.get("voucher_type"), args.get("from_date"), args.get("to_date"),
                args.get("party_type"), args.get("limit", 50), args.get("filters"),
            )
            return json.dumps(data, ensure_ascii=False, default=str)
        elif name == "erp_aggregate":
            data = erp_aggregate(
                args["doctype"], args["group_by"],
                args.get("metric", "sum"),
                args.get("field", "base_grand_total"),
                args.get("filters"),
                args.get("order", "desc"),
                args.get("limit", 50),
            )
            return json.dumps(data, ensure_ascii=False, default=str)
        elif name == "erp_run_report":
            data = erp_run_report(args["report_name"], args.get("filters"))
            return json.dumps(data, ensure_ascii=False, default=str)
        elif name == "erp_profit_breakdown":
            data = erp_profit_breakdown(
                args["from_date"], args["to_date"],
                args.get("by", "item_code"),
                args.get("customer"),
                args.get("top", 20),
                args.get("company"),
                args.get("rank_by", "gross_profit"),
                args.get("order", "desc"),
            )
            return json.dumps(data, ensure_ascii=False, default=str)
        else:
            return f"Noma'lum tool: {name}"
    except Exception as e:
        return f"ERP xatosi: {e}"


# ============ Claude uchun tool ta'riflari =====================================
TOOLS = [
    {
        "name": "send_financial_statement",
        "description": (
            "Kompaniyaning BALANS yoki FOYDA-ZARAR (P&L) hisobotini REAL ERPNext ma'lumotidan "
            "(GL Entry) tuzib, OYLIK ustunlarda, kompaniyaning rus tilidagi hisob nomlari bilan "
            "PDF/Excel qilib yuboradi. Foydalanuvchi 'balans', 'balance', 'foyda-zarar', 'pnl', "
            "'P&L' so'rasa — SHU tool'ni ishlat (data ERPNext'dan real olinadi). "
            "document: 'balance' yoki 'pnl'. period_start_date/period_end_date (YYYY-MM-DD) — "
            "oraliq oylarga bo'linadi (har oy alohida ustun). Sana berilmasa, foydalanuvchidan "
            "qaysi davr ekanini SO'RA (masalan oxirgi 4 oy). file_format: 'pdf' (default), 'xlsx', 'csv'. "
            "Eslatma: bu kompaniyada moliyaviy yil sozlamasi muammoli, shuning uchun standart "
            "Balance Sheet/P&L o'rniga AYNAN shu tool ishlatiladi."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "document": {"type": "string", "enum": ["balance", "pnl"],
                             "description": "balance = balans (Баланс), pnl = foyda-zarar (P&L)"},
                "period_start_date": {"type": "string", "description": "Davr boshi YYYY-MM-DD"},
                "period_end_date": {"type": "string", "description": "Davr oxiri YYYY-MM-DD"},
                "file_format": {"type": "string", "enum": ["pdf", "xlsx", "csv"]},
                "caption": {"type": "string", "description": "Telegram izohi (ixtiyoriy)"},
            },
            "required": ["document", "period_start_date", "period_end_date"],
        },
    },
    {
        "name": "erp_list",
        "description": (
            "ERPNext'dan hujjatlar ro'yxatini o'qiydi (faqat o'qish). "
            "Filtr formati: [[\"maydon\",\"operator\",\"qiymat\"], ...]. "
            "Operatorlar: =, !=, >, <, >=, <=, like, in, between, is. "
            "Sana maydonlari: posting_date, creation, modified. "
            "Kim yaratgani: owner (email). Kim o'zgartirgani: modified_by."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "doctype": {"type": "string",
                            "description": "Masalan: Sales Invoice, Delivery Note, Purchase Receipt, Purchase Invoice, Payment Entry"},
                "filters": {"type": "array", "items": {"type": "array"},
                            "description": "Filtrlar ro'yxati"},
                "fields": {"type": "array", "items": {"type": "string"},
                           "description": "Olinadigan maydonlar, masalan [\"name\",\"customer\",\"grand_total\",\"owner\",\"creation\"]"},
                "limit": {"type": "integer", "description": "Maksimal qatorlar (default 50)"},
                "order_by": {"type": "string", "description": "Masalan: 'creation desc'"},
            },
            "required": ["doctype"],
        },
    },
    {
        "name": "erp_get_doc",
        "description": "Bitta ERPNext hujjatining to'liq tafsilotini o'qiydi (faqat o'qish).",
        "input_schema": {
            "type": "object",
            "properties": {
                "doctype": {"type": "string"},
                "name": {"type": "string", "description": "Hujjat ID/nomi"},
            },
            "required": ["doctype", "name"],
        },
    },
    {
        "name": "erp_ledger",
        "description": (
            "GL Entry (bosh kitob) — TRANZAKSIYALARNI CHUQUR ko'rsatadi: pul qaysi hisobdan "
            "qaysi hisobga ketdi, KIM bilan (party=mijoz/yetkazib beruvchi/xodim), debet/kredit, "
            "qaysi hujjat orqali. Foydalanuvchi 'bu to'lov nima', 'pul qayerdan qayerga ketdi', "
            "'falon mijoz/supplier bilan harakatlar', 'qaysi schotlar ishladi', 'kimga berdik/kimdan "
            "oldik', 'bu hujjatning provodkasi' desa SHU tool'ni ishlat. "
            "Har posting: account (hisob), against (qarama-qarshi hisob), party (taraf), debit, credit, "
            "voucher_type/voucher_no (manba hujjat), against_voucher (bog'liq invoys), remarks. "
            "voucher_no berilsa — aynan o'sha hujjatning ikki tomonlama provodkasi chiqadi."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "party": {"type": "string", "description": "Mijoz/yetkazib beruvchi/xodim nomi"},
                "party_type": {"type": "string", "description": "Customer / Supplier / Employee"},
                "account": {"type": "string", "description": "Hisob nomi yoki qismi (masalan 'Банк', '1311')"},
                "voucher_no": {"type": "string", "description": "Aniq hujjat raqami (to'lov/invoys)"},
                "voucher_type": {"type": "string", "description": "Payment Entry / Sales Invoice / Journal Entry ..."},
                "from_date": {"type": "string", "description": "YYYY-MM-DD"},
                "to_date": {"type": "string", "description": "YYYY-MM-DD"},
                "limit": {"type": "integer", "description": "Maks qatorlar (default 50, maks 200)"},
            },
        },
    },
    {
        "name": "erp_aggregate",
        "description": (
            "Hujjatlarni biror maydon bo'yicha GURUHLAB, yig'indi/son/o'rtachasini "
            "hisoblaydi (server tomonida — tez va aniq). Tahlil savollari uchun shuni ishlat. "
            "Misol: mijoz bo'yicha bu oygi umumiy sotuv => doctype='Sales Invoice', "
            "group_by='customer', metric='sum', field='base_grand_total', "
            "filters=[['docstatus','=',1],['posting_date','between',['2026-05-01','2026-05-31']]]. "
            "Oylik trendni topish uchun har oy uchun alohida chaqir (sana filtrini o'zgartirib)."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "doctype": {"type": "string"},
                "group_by": {"type": "string", "description": "Guruhlash maydoni, masalan 'customer', 'supplier', 'item_code'"},
                "metric": {"type": "string", "enum": ["sum", "count", "avg", "max", "min"]},
                "field": {"type": "string", "description": "Hisoblanadigan maydon, masalan 'base_grand_total' (count uchun kerak emas)"},
                "filters": {"type": "array", "items": {"type": "array"}},
                "order": {"type": "string", "enum": ["desc", "asc"]},
                "limit": {"type": "integer"},
            },
            "required": ["doctype", "group_by"],
        },
    },
    {
        "name": "erp_run_report",
        "description": (
            "ERPNext'ning tayyor hisobotini ishga tushiradi. REAL FOYDA uchun "
            "report_name='Gross Profit' ishlat (mijoz bo'yicha foyda: filters ichida "
            "{'group_by':'Customer','company':..,'from_date':..,'to_date':..}). "
            "DIQQAT: foyda faqat ERPNext'da tovar tannarxi kiritilgan bo'lsa to'g'ri chiqadi."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "report_name": {"type": "string"},
                "filters": {"type": "object", "description": "Hisobot filtrlari (obyekt ko'rinishida)"},
            },
            "required": ["report_name"],
        },
    },
    {
        "name": "erp_profit_breakdown",
        "description": (
            "Tovar/mijoz kesimida SOTUV va FOYDANI hisoblaydi. TOVAR darajasidagi deyarli "
            "barcha savollar uchun shu tool: 'Top 10 mahsulot', 'eng ko'p/kam sotilgan', "
            "'qaysi tovardan ko'p foyda', 'falon mijozning qaysi tovari'. "
            "rank_by bilan saralash o'zgaradi: 'revenue' (savdo/tushum bo'yicha — 'ko'p sotilgan'), "
            "'qty' (miqdor bo'yicha), 'gross_profit' (foyda bo'yicha). "
            "order='asc' => 'eng kam sotilgan'. by='customer' => mijozlar reytingi; "
            "customer='Nomi' => faqat shu mijoz. Sana oralig'i shart. "
            "DIQQAT: '...Item' child table'ni erp_aggregate bilan guruhlama — shu tool'dan foydalan."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "from_date": {"type": "string", "description": "YYYY-MM-DD"},
                "to_date": {"type": "string", "description": "YYYY-MM-DD"},
                "by": {"type": "string", "enum": ["item_code", "customer", "customer_item"]},
                "customer": {"type": "string", "description": "Faqat shu mijoz bilan cheklash (ixtiyoriy)"},
                "rank_by": {"type": "string", "enum": ["gross_profit", "revenue", "qty"],
                            "description": "Saralash mezoni: revenue=savdo, qty=miqdor, gross_profit=foyda"},
                "order": {"type": "string", "enum": ["desc", "asc"],
                          "description": "desc=ko'pdan kamga (top), asc=kamdan ko'pga (eng kam)"},
                "top": {"type": "integer", "description": "Nechta natija (default 20)"},
                "company": {"type": "string", "description": "Kompaniya nomi (sozlamada bo'lsa shart emas)"},
            },
            "required": ["from_date", "to_date"],
        },
    },
    {
        "name": "get_report_data",
        "description": (
            "ERPNext hisobotining RAQAMLARINI o'qiydi (fayl yubormaydi). Hisobotni TAHLIL "
            "qilish, izohlash yoki maslahat berish uchun SHUNI ishlat — raqamlar senga matn "
            "bo'lib qaytadi. report_name: 'Balance Sheet', 'Profit and Loss Statement', 'Cash Flow' "
            "(va boshqalar). Bu 3 hisobot uchun filters'ga faqat period_start_date va "
            "period_end_date (YYYY-MM-DD) ber — qolganini bot avtomatik qo'yadi."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "report_name": {"type": "string"},
                "filters": {"type": "object", "description": "period_start_date, period_end_date va h.k."},
            },
            "required": ["report_name"],
        },
    },
    {
        "name": "send_report_file",
        "description": (
            "ERPNext hisobotini HAQIQIY FAYL (Excel/CSV/PDF) qilib foydalanuvchiga "
            "Telegram orqali yuboradi. Foydalanuvchi 'PDF qilib jo'nat', 'Excel/fayl "
            "qilib ber', 'yuklab ber' desa SHU tool'ni ishlat — matn jadval emas, fayl yubor. "
            "report_name — ERPNext hisobot nomi (masalan: 'Balance Sheet', 'Profit and Loss "
            "Statement', 'General Ledger', 'Accounts Receivable', 'Trial Balance'). "
            "file_format: 'pdf' (chiroyli, chop etishga tayyor), 'xlsx' (Excel, tahrirlash uchun) "
            "yoki 'csv'. Foydalanuvchi formatni aytmasa: PDF so'ralganda 'pdf', aks holda 'xlsx'. "
            "filters — hisobot filtrlari OBYEKT ko'rinishida.\n"
            "MUHIM: Balance Sheet, 'Profit and Loss Statement', 'Cash Flow' uchun SEN faqat "
            "filters'ga period_start_date va period_end_date (YYYY-MM-DD) ber. filter_based_on, "
            "periodicity (Monthly), accumulated_values ni bot O'ZI to'g'ri qo'yadi — ularni yozma. "
            "Sana berilmasa foydalanuvchidan qaysi davr ekanini avval SO'RA."
        ),
        "input_schema": {
            "type": "object",
            "properties": {
                "report_name": {"type": "string", "description": "ERPNext hisobot nomi"},
                "filters": {"type": "object", "description": "Hisobot filtrlari (obyekt)"},
                "file_format": {"type": "string", "enum": ["pdf", "xlsx", "csv"]},
                "filename": {"type": "string", "description": "Fayl nomi (kengaytmasiz, ixtiyoriy)"},
                "title": {"type": "string", "description": "Fayl ichidagi sarlavha (ixtiyoriy)"},
                "caption": {"type": "string", "description": "Telegram'dagi fayl izohi (ixtiyoriy)"},
            },
            "required": ["report_name"],
        },
        # Prompt caching: oxirgi tool'gacha bo'lgan BARCHA tool ta'riflari keshlanadi
        # (har so'rovda qaytadan yuborilmaydi => token va pul tejaladi).
        "cache_control": {"type": "ephemeral"},
    },
]

_WEEKDAYS_UZ = ["Dushanba", "Seshanba", "Chorshanba", "Payshanba", "Juma", "Shanba", "Yakshanba"]


def _system_blocks():
    """Claude uchun system promptni ikki blokda qaytaradi:
      1) STATIK ko'rsatmalar — prompt caching bilan keshlanadi (token tejaladi);
      2) bugungi SANA bloki — har kuni o'zgaradi, keshlanmaydi.
    COMPANY va sana har so'rovda joriy qiymatdan olinadi (autodetect'dan keyin ham to'g'ri)."""
    company_filter = (f'["company","=","{COMPANY}"]' if COMPANY
                      else 'kerak emas (hamma kompaniya)')
    cur = CURRENCY or "USD"
    # .format() ISHLATMAYMIZ — matnda JSON misollaridagi { } qavslar bor.
    static = (_SYSTEM_BODY
              .replace("__COMPANY_FILTER__", company_filter)
              .replace("__CURRENCY__", cur))
    n = _now()
    date_block = (
        f"MUHIM — BUGUNGI SANA: {n:%Y-%m-%d} ({_WEEKDAYS_UZ[n.weekday()]}), "
        f"vaqt mintaqasi {TIMEZONE}.\n"
        "'bugun', 'kecha', 'shu hafta', 'bu oy', 'o'tgan oy', 'shu yil' kabi nisbiy "
        "iboralarni AYNAN shu sanaga nisbatan hisobla. Hech qachon sanani taxmin qilma — "
        "agar davr noaniq bo'lsa, foydalanuvchidan aniqlashtirib so'ra."
    )
    return [
        {"type": "text", "text": static, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": date_block},
    ]


_SYSTEM_BODY = """Sen — Ruxsora Shirinliklari kompaniyasining ERPNext yordamchisisan, o'zbek tilida
javob berasan. Foydalanuvchi — BIZNES EGASI / rahbar, moliyachi yoki buxgalter EMAS. Shuning
uchun SODDA, aniq va tushunarli gapir: buxgalteriya atamalari (debet/kredit, provodka, schot
raqami) o'rniga oddiy tilda "pul kirdi / pul chiqdi / kimga / nima uchun" deb tushuntir.
Kerakli ma'lumotni tool'lar orqali ERPNext'dan olib, qisqa va aniq javob ber.

═══ PUL BIRLIGI — JUDA MUHIM (eng ko'p chalkashlik shu yerda) ═══
- Kompaniyaning hisob valyutasi: __CURRENCY__. ERPNext'dan kelgan BARCHA summalar (balans,
  P&L, GL, invoys, to'lov, qoldiq, foyda) AYNAN shu valyutada — __CURRENCY__.
- Summani DOIM to'g'ri belgila. __CURRENCY__ = USD bo'lsa "$" yoki "USD" yoz.
  HECH QACHON summalarni "so'm" deb yozma — raqamlar dollarda, "so'm" desang foydalanuvchi
  butunlay yanglishadi (masalan 5 357 USD ni "5 357 so'm" desang — 800 barobar xato).
- "ming", "mln", "milliard" so'zlarini O'ZINGDAN QO'SHMA. Raqamni borligicha yoz, faqat
  minglikni bo'sh joy bilan ajrat: "354 862 USD" (NE "354 862 ming so'm").
- Bir to'lov boshqa valyutada kiritilgan bo'lishi mumkin (izoh/remarks'da "UZS 100000000").
  Bu — original summa. Lekin buxgalteriya (debit/credit) summasi __CURRENCY__ da. Foydalanuvchiga
  ASOSIY summani (__CURRENCY__) ber, originalni faqat qavsda eslat:
  "8 183 USD (mijoz 100 000 000 so'm to'lagan)". BIR summani ikki marta ikki xil ko'rsatma!

Foydali bilimlar:
- Faqat tasdiqlangan hujjatlar uchun filtrga ["docstatus","=",1] qo'sh.
- "Invoys qilinmagan Delivery Note" = Delivery Note, status="To Bill".
- "Invoys qilinmagan Purchase Receipt" = Purchase Receipt, status="To Bill".
- "Kim kiritgan/yaratgan" = 'owner' maydoni (email). 'creation' = yaratilgan vaqt.
- Naqd kassa = Payment Entry, mode_of_payment="Cash" (kompaniyada nomi boshqa bo'lishi mumkin).
- Kompaniya filtri: __COMPANY_FILTER__.
- Sanani 'YYYY-MM-DD' formatida ishlat. Oraliq uchun 'between' operatori: ["posting_date","between",["2026-05-01","2026-05-31"]].

BALANS / FOYDA-ZARAR (P&L) HISOBOTLARI:
- Foydalanuvchi "balans", "balance", "foyda-zarar", "pnl", "P&L" so'rasa — send_financial_statement
  tool'ini ishlat (document="balance" yoki "pnl"). Data REAL ERPNext'dan (GL Entry) olinadi,
  oylik ustunlarda, kompaniyaning rus tilidagi hisob nomlari bilan chiqadi.
- period_start_date/period_end_date BER. Sana aytilmasa, foydalanuvchidan qaysi davr ekanini
  SO'RA (masalan "oxirgi 4 oy" => bugungi oydan orqaga 4 oy).
- Bu kompaniyada moliyaviy yil sozlamasi muammoli — shuning uchun balans/pnl uchun DOIM
  send_financial_statement ishlat, send_report_file("Balance Sheet"/"Profit and Loss") ISHLATMA.
- Boshqa hisobotlar (General Ledger, Trial Balance, Accounts Receivable, Cash Flow...) uchun
  esa send_report_file yoki get_report_data ishlatishda davom et.

============ CHUQUR (BOY) JAVOB — JUDA MUHIM ============
Foydalanuvchi biror davr uchun sotuv/xarid/to'lov/harakat so'rasa, QURUQ RO'YXAT bilan
cheklanma. O'zing kerakli tool'larni ketma-ket chaqirib, TO'LIQ surat ber. Har doim:
- JAMI summa + nechta hujjat ("Bugun 12 ta sotuv, jami 5 320 USD").
- ENG KATTA(lar): eng baland 3-5 tasi — kim va qancha.
- KIM bilan: mijoz/yetkazib beruvchi bo'yicha taqsimot (kim ko'p oldi yoki berdi).
- PUL QAYERGA/QAYERDAN: pul naqdmi, bankkami tushdi yoki qarzga ketdimi. To'lov bo'lsa
  erp_ledger (account → against, party) bilan "qaysi hisobdan qaysi hisobga" ni ko'rsat.
- TO'LOV HOLATI: sotuv to'langanmi yoki qarz (outstanding_amount > 0).
- Oxirida bitta jumlali XULOSA ("Eng ko'p X mijoz oldi, ammo hali to'lamagan").
Bu chuqurlik FAQAT sotuvga emas — xarid, to'lov, kassa, ombor, ishlab chiqarish, foydalanuvchi
harakatlariga ham tegishli (har birida: jami, top, kim, qayerga/qayerdan, holat, xulosa).
Ortiqcha savol berma — ma'lumotni o'zing yig'ib, tayyor boy javob ber.
Masalan "bugungi sotuvlar":
  1) erp_list Sales Invoice (bugun, docstatus=1) fields=customer, grand_total, outstanding_amount, status.
  2) Jami + nechta; eng katta 3-5 mijoz/invoice; to'langan/qarz holati.
  3) Kerak bo'lsa erp_ledger bilan pul qayerga tushganini qo'sh.
  4) Qisqa xulosa.

============ BILIMLAR BAZASI: SAVOL TURI => QAYERDAN OLINADI ============
Quyidagi yo'riqnomaga qat'iy amal qil. Sotuv/xarid summalari uchun DOIM ["docstatus","=",1]
filtrini qo'sh. Sana oralig'i uchun ["posting_date","between",["BOSH","OXIR"]].

— SAVDO (Sales) —
- "Bu oy savdo qancha / jami sotuv" => erp_aggregate, doctype="Sales Invoice",
  metric="sum", field="base_grand_total", group_by="company" (yoki istalgan), posting_date filtri.
  (group_by shart; umumiy yig'indi uchun group_by="company" qo'yib, natijani qo'sh.)
- "O'tgan oy bilan solishtir" => shu oy uchun bitta, o'tgan oy uchun ikkinchi erp_aggregate;
  keyin farq va foiz o'zgarishini hisobla.
- "Top 10 mahsulot / eng ko'p sotilgan" => erp_profit_breakdown, by="item_code",
  rank_by="revenue" (yoki "qty"), order="desc", top=10.
- "Eng kam sotilgan mahsulotlar" => erp_profit_breakdown, by="item_code", rank_by="qty",
  order="asc". (Eslatma: bu faqat SOTILGAN tovarlar ichida eng kami; umuman sotilmaganlar
  uchun Stock hisoboti kerak — buni aytib qo'y.)
- "Qaysi filial ko'proq sotmoqda" => erp_aggregate, doctype="Sales Invoice",
  group_by="branch" (ishlamasa group_by="territory" yoki "cost_center" yoki "set_warehouse").
- "Kunlik savdo trendi" => erp_aggregate, doctype="Sales Invoice", group_by="posting_date",
  metric="sum", field="base_grand_total", order="asc", posting_date filtri.
- "Haftalik savdo trendi" => avval kunlik olib (yuqoridagidek), keyin haftalarga o'zing yig'.

— OMBOR (Stock) —
- "Joriy qoldiq / eng ko'p qoldiq qaysi mahsulotda" => erp_list, doctype="Bin",
  fields=["item_code","warehouse","actual_qty","valuation_rate"], order_by="actual_qty desc".
  Yoki to'liq surat uchun erp_run_report("Stock Balance", {"company":..,"from_date":..,"to_date":..}).
- "Qaysi mahsulotlar tugab qolmoqda / safety stockdan past" =>
  erp_run_report("Item Shortage Report") yoki Bin'da actual_qty kichik bo'lganlarini ko'r.
- "Slow moving / dead stock / uzoq turgan tovar" => erp_run_report("Stock Ageing",
  {"company":..,"to_date":..}); yoshi katta (uzoq harakatsiz) tovarlar = slow/dead stock.
- "Inventory turnover" => standart bitta hisobot yo'q. Stock Balance (zaxira qiymati) va
  shu davr sotilgan miqdor/COGS asosida taxminiy hisobla va metodikani tushuntir.

— XARID (Purchase) —
- "Bu oy qancha xarid" => erp_aggregate, doctype="Purchase Invoice", metric="sum",
  field="base_grand_total", group_by="company", posting_date filtri (docstatus=1).
- "Eng ko'p xarid qilingan mahsulotlar" => erp_run_report("Item-wise Purchase Register",
  {"company":..,"from_date":..,"to_date":..}); item bo'yicha qty/amount natijasini saralab ber.
- "Yetkazib beruvchi bo'yicha xarid" => erp_aggregate, doctype="Purchase Invoice",
  group_by="supplier", metric="sum", field="base_grand_total".
- "Xarid narxlari o'zgarishi" => Item-wise Purchase Register'dagi rate ustunini davrlar
  bo'yicha solishtir, yoki erp_list "Item Price" (buying narxlari).

— TRANZAKSIYALAR / PUL OQIMI (chuqur tafsilot) —
- "Bu to'lov/hujjat nima edi", "pul qayerdan qayerga ketdi", "qaysi schotlar (hisoblar)
  ishladi", "provodka" => erp_ledger(voucher_no="HUJJAT-RAQAMI"). Natijada ikki tomonlama
  yozuv chiqadi: account (debet/kredit qilingan hisob) va against (qarama-qarshi hisob).
  Tushuntir: "X hisobdan Y hisobga Z so'm o'tdi".
- "Falon mijoz/yetkazib beruvchi bilan harakatlar / hisob-kitob" => erp_ledger(party="Nomi",
  from_date, to_date). debet/kredit bo'yicha kim kimga qarzdorligini ko'rsat.
- "Kimga sotdik / kim oldi" => Sales Invoice (customer). "Kimdan oldik / kimga to'ladik" =>
  Purchase Invoice (supplier) yoki Payment Entry. To'lov tafsiloti uchun erp_ledger yoki
  erp_get_doc("Payment Entry", nomi): paid_from = pul chiqgan hisob, paid_to = pul kirgan hisob,
  party = qarshi taraf.
- "Bank/kassa harakati" => erp_ledger(account="Банк") yoki account="Наличные", sana oralig'i bilan.
- Provodkani tushuntirganda DOIM: sana, summa, qaysi hisobdan→qaysi hisobga, party (kim),
  manba hujjat (voucher_no) va izoh (remarks) ni ber. Raqamlarni o'qiladigan qil.

— MOLIYA (ERPNext) —
- "Debitor qarzdorlik / mijoz qarzlari" => erp_run_report("Accounts Receivable",
  {"company":..,"report_date":"<bugun yoki so'ralgan sana>"}).
- "Kreditor qarzdorlik" => erp_run_report("Accounts Payable", {"company":..,"report_date":..}).
- "Cash flow / pul oqimi" => get_report_data("Cash Flow", period_start_date/period_end_date bilan).
- "Outstanding / to'lanmagan fakturalar" => erp_list, doctype="Sales Invoice",
  filters=[["docstatus","=",1],["status","in",["Unpaid","Overdue","Partly Paid"]]],
  fields=["name","customer","grand_total","outstanding_amount","due_date","status"].

— FOYDALANUVCHI FAOLIYATI —
- "Bugun/kecha kim invoice yaratdi" => erp_list, doctype="Sales Invoice",
  filters=[["creation","between",["<sana> 00:00:00","<sana> 23:59:59"]]],
  fields=["name","owner","creation","customer","grand_total"]. owner = yaratgan odam (email).
- "Qaysi operator eng ko'p hujjat kiritgan / user activity ranking" => erp_aggregate,
  kerakli doctype, group_by="owner", metric="count" (creation sanasi filtri bilan).
- "Kim stock entry / sales order yaratgan" => erp_list, doctype="Stock Entry" yoki
  "Sales Order", fields=["name","owner","creation"], creation filtri bilan.
- "Kim oxirgi 7 kunda tizimga kirmagan" => erp_list, doctype="User",
  fields=["name","full_name","last_active","last_login","enabled"],
  filters=[["enabled","=",1],["last_active","<","<bugun-7kun> 00:00:00"]].

— HR (agar ERPNext HR moduli yoqilgan bo'lsa) —
- "Davomat" => erp_list, doctype="Attendance",
  fields=["employee_name","attendance_date","status","working_hours"], attendance_date filtri.
- "Kechikishlar" => Attendance filters=[["late_entry","=",1]] yoki "Employee Checkin".
- "Ta'tillar" => erp_list, doctype="Leave Application",
  fields=["employee_name","leave_type","from_date","to_date","status"].
- "Ish soatlari" => Attendance "working_hours" yoki erp_list "Timesheet".
- Agar tool "DocType ... not found" yoki shunga o'xshash xato qaytarsa => HR moduli
  yoqilmagan bo'lishi mumkin; buni foydalanuvchiga sodda tilda ayt.

UMUMIY QOIDALAR:
- "REAL FOYDA" so'ralsa => erp_profit_breakdown yoki erp_run_report("Gross Profit").
  Tannarx (valuation) kiritilmagan bo'lsa foyda bo'sh chiqadi — buni ayt va tushum
  (revenue) bo'yicha tahlil taklif qil. Tushum ≠ foyda ekanini aniq tushuntir.
- "Falon mijozning qaysi tovari" => erp_profit_breakdown, customer="Mijoz", by="item_code".
- Maydon nomi noto'g'ri bo'lib xato kelsa, muqobil maydonni sina (masalan branch↔territory).
- "...Item" child table'ni erp_aggregate bilan GURUHLAMA — tovar tahlili uchun erp_profit_breakdown.
- Oyma-oy yoki "oldin va hozir" solishtiruvda: har davr uchun alohida chaqir, so'ng farq va
  foiz o'zgarishini ko'rsat (masalan "aprel 50 mln → may 18 mln, 64% pasaygan").
- So'ralgan ma'lumot tizimda bo'lmasa yoki tool xato qaytarsa — TO'QIMA, ochiq ayt.

FAYL (PDF/Excel) YUBORISH:
- Foydalanuvchi hisobotni "PDF qilib jo'nat", "Excel/fayl qilib ber", "yuklab ber" desa —
  send_report_file tool'ini ishlat. SEN FAYL YUBORA OLASAN — "imkonim yo'q" DEMA.
- 3 ta asosiy moliyaviy hisobot: report_name="Balance Sheet" (balans),
  "Profit and Loss Statement" (foyda-zarar), "Cash Flow" (pul oqimi).
- Bu 3 hisobot uchun SEN faqat sana berasan — filters'ga period_start_date va
  period_end_date (YYYY-MM-DD). filter_based_on/periodicity/accumulated ni bot o'zi qo'yadi.
    * Balance Sheet "as on" hisobot: period_start_date = moliyaviy yil boshi (2026-01-01),
      period_end_date = so'ralgan sana (masalan 2026-04-30).
    * Profit and Loss / Cash Flow: period_start_date va period_end_date = so'ralgan davr
      boshi va oxiri (masalan "mart-may" => 2026-03-01 ... 2026-05-31; "aprel" => 2026-04-01 ... 2026-04-30).
- Format aytilmasa: "PDF" so'ralsa file_format="pdf", aks holda "xlsx" (Excel).
- Sana berilmasa, foydalanuvchidan qaysi davr ekanini avval SO'RA.
- Fayl ketgach qisqa tasdiq yoz (masalan "Balance Sheet PDF faylini yubordim ✅").

HISOBOTNI TAHLIL QILISH (CFO / moliyachi sifatida):
- Foydalanuvchi "tahlil qil", "CFO/moliyachi sifatida bahola", "nima qilish kerak",
  "izohlab ber" desa — kerakli hisobot raqamlarini get_report_data bilan O'ZING ol va tahlil qil.
- HECH QACHON foydalanuvchidan raqamlarni qo'lda kiritishini SO'RAMA. "PDF/Excel faylni
  o'qiy olmayman" deb uzr ham so'rama — ma'lumot senda get_report_data orqali bor.
  (Fayl yuborgan bo'lsang, uning ma'lumoti tool natijasida senga allaqachon qaytgan.)
- Tahlilda quyidagilarni yorit: asosiy ko'rsatkichlar (aktiv/passiv/sarmoya yoki
  tushum/xarajat/sof foyda yoki sof pul oqimi), muhim nisbatlar (joriy likvidlik,
  foyda marjasi, qarz/sarmoya), oylar bo'yicha TREND (o'sish/pasayish %), e'tibor
  beriladigan xavflar, va 3-5 ta ANIQ amaliy tavsiya. Raqamlarni o'qiladigan qil.
- Tahlil uchun bir nechta hisobotni birga olib solishtirsang bo'ladi (masalan P&L + Cash Flow).

JAVOB KO'RINISHI (UI) — foydalanuvchi BIZNES EGASI, Telegram'da o'qiydi:
- Telegram oddiy MATN ko'rsatadi (markdown emas). '**', '#', '|' jadval, '```' belgilarini
  ISHLATMA — xunuk ko'rinadi. Toza matn yoz. Sarlavhaga emoji + BOSH HARF mumkin ("📊 SAVDO").
- Avval BITTA jumlada asosiy javob/xulosa, keyin kerak bo'lsa tafsilot. Qisqa va lo'nda.
- Ro'yxat uchun "• " yoki "1) 2) 3)". Raqamni minglik bo'sh joy bilan: 1 234 567 + VALYUTA.
  O'sish/pasayishni "▲ 12%" / "▼ 8%", sanani "17.06.2026" ko'rinishida.
- Uzun ro'yxat (8-10 dan ko'p) bo'lsa eng muhim 10 tasini ko'rsat, "to'liqini Excel qilib
  beraymi?" deb taklif qil.
- JADVAL (| ustun |) MUTLAQO ISHLATMA — Telegram'da quvurlar va yulduzchalar bo'lib chiqadi.
  Oylarni solishtirganda har oyni ALOHIDA QATORDA ber. Masalan TO'G'RI ko'rinish:
    📊 Yanvar–May 2026 (P&L)
    • Yanvar: savdo 117 890, foyda 8 361 (7.1%)
    • Fevral: savdo 130 003, foyda 2 317 (1.8%)
    • May: savdo 105 940, foyda −5 773 (ZARAR ▼)
  YOMON ko'rinish (ishlatma): "| Oy | Savdo | Foyda |" va "**qalin**" belgilar.

SODDA TIL (buxgalter emas, biznes egasi uchun):
- Buxgalteriya atamalarini sodda tilga o'gir: "debet/kredit" o'rniga "pul kirdi/chiqdi" yoki
  "qarz oshdi/kamaydi"; "provodka" o'rniga "bu pul qayerdan qayerga o'tgani".
- Hisob (schot) raqamlarini (masalan "5215", "1220 - Банк р/c") odatda KO'RSATMA — uning
  ODDIY MA'NOSINI yoz: "Bank hisobidan", "Soliq xarajatiga", "Mijozlar qarzi". Faqat
  foydalanuvchi aniq "schot/hisob raqami" so'rasa kodni ber.
- Rus tilidagi hisob nomlarini o'zbekchaga tushuntir (Аренда помещение = bino ijarasi,
  Дилерский бонусы = dilerlik bonusi, Дебиторская задолженность = mijozlar qarzi).

ATAMALAR — DOIM bir xil ta'rifga amal qil (javoblar ziddiyatli bo'lmasin):
- "Savdo/sotuv" = Sales Invoice (docstatus=1). "Xarid" = Purchase Invoice.
- "To'lov" = Payment Entry. "Naqd/kassa" = naqd pul (Наличные / mode_of_payment Cash).
  "Bank" = bank hisobi (Банк р/c). Foydalanuvchi "kassa/to'lovlar" desa — naqd VA bank
  to'lovlarini ALOHIDA ko'rsat (to'liq surat), shunda javob ziddiyatsiz bo'ladi.
- Bir savol qayta berilsa, AVVALGI javob bilan mos kelsin — filtr/davrni o'zboshimcha
  o'zgartirib boshqa raqam berma. "Kechagi to'lovlar" har safar bir xil chiqishi shart.

XATO YUZ BERSA:
- Foydalanuvchiga HECH QACHON xom xatoni (kod 417, "EXPECTATION FAILED", maydon nomlari,
  stack-trace) ko'rsatma — bu chalkashtiradi. Bitta sodda jumlada ayt va DARHOL boshqa usul
  bilan urin (masalan Journal Entry o'rniga erp_ledger, yoki boshqa maydon). Texnik tafsilot
  berma va "kerakmi?" deb so'rama — muammoni o'zing yech, faqat natijani ber.

DAVR (sana) berilmaganda:
- Hisobot/tahlil so'ralsa-yu davr aytilmasa, har safar SO'RAB o'tirma. Eng mantiqiy davrni
  (odatda shu oy yoki oxirgi to'liq oy) O'ZING tanla, qaysi davr olganingni AYTIB qo'y va
  "boshqa davr kerak bo'lsa ayting" deb qo'sh. Faqat haqiqatan noaniq bo'lsagina so'ra.

Hech qachon ma'lumotni o'zingdan to'qima — faqat tool natijasiga asoslanib javob ber."""


# ============ Claude API bilan agentik tsikl =================================
# Qayta urinish sozlamalari (rate limit / vaqtinchalik xatolar uchun)
ANTHROPIC_MAX_RETRIES = 6     # ko'pi bilan necha marta qayta urinish
ANTHROPIC_BACKOFF_CAP = 60    # ikki urinish orasidagi maksimal kutish (soniya)


def _anthropic_post(payload):
    """Anthropic'ga POST yuboradi. 429 (rate limit), 529 (overloaded) va 5xx
    xatolarida avtomatik qayta uradi — Anthropic bergan 'retry-after' sarlavhasini
    hisobga olib, eksponensial backoff bilan. Foydalanuvchi xatoni ko'rmaydi."""
    headers = {
        "x-api-key": ANTHROPIC_KEY,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json",
    }
    last_exc = None
    for attempt in range(ANTHROPIC_MAX_RETRIES + 1):
        try:
            r = requests.post(
                "https://api.anthropic.com/v1/messages",
                headers=headers, json=payload, timeout=120,
            )
            # Qayta urinsa bo'ladigan xatolar: 429 va 5xx (529 ham shu ichida)
            if r.status_code == 429 or r.status_code >= 500:
                if attempt < ANTHROPIC_MAX_RETRIES:
                    wait = _retry_wait(r, attempt)
                    print(f"Anthropic {r.status_code} — {wait:.0f}s kutib qayta urinaman "
                          f"({attempt + 1}/{ANTHROPIC_MAX_RETRIES})")
                    time.sleep(wait)
                    continue
            r.raise_for_status()
            return r.json()
        except requests.exceptions.RequestException as e:
            # Tarmoq/timeout xatosi — bularni ham qayta urinamiz
            last_exc = e
            if attempt < ANTHROPIC_MAX_RETRIES:
                wait = min(ANTHROPIC_BACKOFF_CAP, 2 ** attempt)
                print(f"Anthropic tarmoq xatosi ({e}) — {wait}s kutib qayta urinaman "
                      f"({attempt + 1}/{ANTHROPIC_MAX_RETRIES})")
                time.sleep(wait)
                continue
            raise
    if last_exc:
        raise last_exc
    raise RuntimeError("Anthropic'ga ulanib bo'lmadi (qayta urinishlar tugadi).")


def _retry_wait(resp, attempt):
    """Keyingi urinishgacha necha soniya kutishni hisoblaydi.
    Avval server bergan 'retry-after' sarlavhasiga ishonamiz, bo'lmasa
    eksponensial backoff (2,4,8,...) — maksimal ANTHROPIC_BACKOFF_CAP gacha."""
    ra = resp.headers.get("retry-after")
    if ra:
        try:
            return min(ANTHROPIC_BACKOFF_CAP, float(ra))
        except ValueError:
            pass
    return min(ANTHROPIC_BACKOFF_CAP, 2 ** attempt)


# Tahlil/CFO talab qiladigan savollar — bularga kuchliroq model (MODEL_SMART) ishlatiladi.
_SMART_KEYWORDS = (
    "tahlil", "analiz", "cfo", "moliyachi", "baho", "izoh", "tushuntir", "nima qilish",
    "tavsiya", "maslahat", "trend", "solishtir", "prognoz", "strategiya", "sabab",
    "nega", "nima uchun", "xulosa", "dinamika", "o'sish", "pasay",
)


def _pick_model(text):
    """So'rovga qarab modelni tanlaydi: oddiy savol => MODEL (arzon),
    tahlil/izoh talab qiladigan savol => MODEL_SMART (kuchliroq, agar sozlangan bo'lsa).
    Shunday qilib token va pul tejaymiz, sifatni esa kerak joyda oshiramiz."""
    if MODEL_SMART:
        t = (text or "").lower()
        if any(k in t for k in _SMART_KEYWORDS):
            return MODEL_SMART
    return MODEL


def _safe_truncate_result(out, limit=50000):
    """Tool natijasi juda uzun bo'lsa, kesib, ochiq belgi qo'shamiz — Claude
    natija qisqartirilganini biladi va noto'g'ri xulosa chiqarmaydi."""
    if len(out) <= limit:
        return out
    return out[:limit] + "\n…[natija juda uzun bo'lgani uchun qisqartirildi]"


def ask_claude(history, chat_id=None, model=None):
    """history — Anthropic messages ro'yxati. Tool'larni bajarib, yakuniy javobni qaytaradi.
    chat_id — fayl yuboradigan tool'lar uchun. model — tanlangan model (yo'q bo'lsa MODEL)."""
    use_model = model or MODEL
    for _ in range(8):   # ko'pi bilan 8 marta tool chaqirishga ruxsat
        data = _anthropic_post({
            "model": use_model,
            "max_tokens": 4096,        # uzun CFO tahlili o'rtada kesilmasligi uchun
            "temperature": 0,          # faktik/moliyaviy javob — barqaror va aniq bo'lsin
            "system": _system_blocks(),  # statik qism keshlanadi + bugungi sana bloki
            "tools": TOOLS,
            "messages": history,
        })
        history.append({"role": "assistant", "content": data["content"]})

        if data.get("stop_reason") == "tool_use":
            # Claude bir yoki bir nechta tool chaqirdi — bajaramiz
            results = []
            for block in data["content"]:
                if block["type"] == "tool_use":
                    out = run_tool(block["name"], block["input"], chat_id)
                    results.append({
                        "type": "tool_result",
                        "tool_use_id": block["id"],
                        "content": _safe_truncate_result(out),
                    })
            history.append({"role": "user", "content": results})
            continue   # natija bilan Claude'ga qайта murojaat

        # Tool kerak emas — yakuniy matnli javob
        return "".join(b["text"] for b in data["content"] if b["type"] == "text")

    return "Kechirasiz, javobni tayyorlay olmadim (juda ko'p urinish)."


# ============ Telegram long-polling ==========================================
SESSIONS = {}        # chat_id -> messages tarixi (oddiy xotira; restartda o'chadi)
SESSION_TS = {}      # chat_id -> oxirgi faollik vaqti (eski sessiyalarni tozalash uchun)
SESSION_LOCKS = {}   # chat_id -> Lock (bitta chatdan ikki xabar bir vaqtda kelsa, ketma-ket)
SESSIONS_GUARD = threading.Lock()  # SESSIONS/SESSION_TS/SESSION_LOCKS lug'atlarini himoya qiladi
SESSION_TTL = 6 * 3600   # 6 soat faolsiz sessiya o'chiriladi
SESSION_MAX = 500        # ko'pi bilan shuncha sessiya saqlanadi (xotira o'smasligi uchun)
ERP_TG_API = f"https://api.telegram.org/bot{TG_TOKEN}"


def _get_chat_lock(chat_id):
    with SESSIONS_GUARD:
        lock = SESSION_LOCKS.get(chat_id)
        if lock is None:
            lock = SESSION_LOCKS[chat_id] = threading.Lock()
        return lock


def _prune_sessions():
    """Eskirgan yoki haddan ko'p sessiyalarni o'chiradi (xotira sızmasligi uchun)."""
    now = time.time()
    with SESSIONS_GUARD:
        old = [cid for cid, ts in SESSION_TS.items() if now - ts > SESSION_TTL]
        for cid in old:
            SESSIONS.pop(cid, None)
            SESSION_TS.pop(cid, None)
            SESSION_LOCKS.pop(cid, None)
        if len(SESSIONS) > SESSION_MAX:
            # eng eski faollikdagilarni o'chiramiz
            for cid, _ in sorted(SESSION_TS.items(), key=lambda kv: kv[1])[:len(SESSIONS) - SESSION_MAX]:
                SESSIONS.pop(cid, None)
                SESSION_TS.pop(cid, None)
                SESSION_LOCKS.pop(cid, None)


def _trim_history(hist, keep=20):
    """Tarixni cheklaydi, LEKIN tool_use/tool_result juftligini buzmaydi.
    Kesish nuqtasini oddiy foydalanuvchi xabaridan (role=user, content=matn)
    boshlanadigan joyga suradi — aks holda Anthropic API 400 xato qaytaradi."""
    if len(hist) <= keep:
        return
    cut = len(hist) - keep
    while cut < len(hist):
        m = hist[cut]
        if m.get("role") == "user" and isinstance(m.get("content"), str):
            break
        cut += 1
    if cut < len(hist):     # xavfsiz nuqta topildi
        del hist[:cut]
    # topilmasa — kesmaymiz (to'g'rilik xotiradan muhimroq)


_MD_HEADING = re.compile(r"^\s{0,3}#{1,6}\s*", re.M)        # ## sarlavha
_MD_HR = re.compile(r"^\s*([-*_])\1{2,}\s*$", re.M)          # --- *** ___ ajratuvchi qator
_MD_TABLE_SEP = re.compile(r"^\s*\|?[\s:|-]*-[\s:|-]*\|?\s*$", re.M)  # |---|---| qator
_MD_BULLET = re.compile(r"^(\s*)[*+]\s+", re.M)             # "* band" -> "• band"


def _strip_markdown(text):
    """Telegram oddiy matn ko'rsatadi — markdown belgilarini (** | --- # `) olib tashlaymiz,
    aks holda foydalanuvchi yulduzcha/quvur/chiziqlarni ko'radi (UI buziladi).
    Model qoidaga rioya qilmasa ham, chiqish DOIM toza bo'ladi."""
    if not text:
        return text
    t = text.replace("`", "")
    t = _MD_HEADING.sub("", t)
    t = _MD_TABLE_SEP.sub("", t)          # jadval ajratuvchi qatorini o'chiramiz
    t = _MD_HR.sub("", t)                 # gorizontal chiziqni o'chiramiz
    out = []
    for line in t.split("\n"):
        if line.count("|") >= 2:          # jadval qatori -> toza matnga aylantiramiz
            cells = [c.strip() for c in line.strip().strip("|").split("|")]
            line = "   ".join(c for c in cells if c)
        out.append(line)
    t = "\n".join(out)
    t = _MD_BULLET.sub(r"\1• ", t)        # "* " / "+ " bandlarni "• " ga
    t = t.replace("**", "").replace("__", "")  # qolgan qalin belgilarini olib tashlaymiz
    t = re.sub(r"\n{3,}", "\n\n", t)      # ortiqcha bo'sh qatorlarni qisqartiramiz
    return t.strip()


def _split_for_telegram(text, limit=3800):
    """Uzun matnni bo'laklarga ajratadi, lekin imkon qadar QATOR chegarasidan kesadi
    (so'z/qator o'rtasidan kesib xabarni buzmaydi)."""
    parts, buf = [], ""
    for line in text.split("\n"):
        if len(line) > limit:                     # juda uzun bitta qator — majburan kesamiz
            if buf:
                parts.append(buf); buf = ""
            for i in range(0, len(line), limit):
                parts.append(line[i:i + limit])
            continue
        if len(buf) + len(line) + 1 > limit:
            parts.append(buf); buf = line
        else:
            buf = f"{buf}\n{line}" if buf else line
    if buf:
        parts.append(buf)
    return parts or [""]


def tg_send(chat_id, text):
    for part in _split_for_telegram(_strip_markdown(text)):
        requests.post(f"{ERP_TG_API}/sendMessage",
                      data={"chat_id": chat_id, "text": part}, timeout=30)


def tg_send_document(chat_id, filename, content, caption=None):
    """Telegram'ga fayl (hujjat) yuboradi. content — bayt-massiv."""
    data = {"chat_id": chat_id}
    if caption:
        data["caption"] = caption[:1024]
    requests.post(f"{ERP_TG_API}/sendDocument",
                  data=data,
                  files={"document": (filename, content)},
                  timeout=120)


def handle_message(msg):
    """Bitta Telegram xabarini to'liq qayta ishlaydi (alohida thread'da chaqiriladi).
    Bitta chatdan ikki xabar bir vaqtda kelsa, per-chat lock ularni ketma-ket bajaradi —
    shunda tarix (history) buzilmaydi va uzun hisobot boshqalarni bloklamaydi."""
    text = msg.get("text")
    chat_id = msg.get("chat", {}).get("id")
    user_id = msg.get("from", {}).get("id")
    if not text or chat_id is None:
        return

    # Xavfsizlik: faqat ruxsat etilgan foydalanuvchilar
    if user_id not in ALLOWED_USERS:
        _log("RAD ETILDI", f"user={user_id}", f"chat={chat_id}")
        tg_send(chat_id, "Kechirasiz, sizga bu botdan foydalanishga ruxsat yo'q.")
        return

    cmd = text.strip()
    if cmd in ("/start", "/help"):
        tg_send(chat_id, "Salom! Men ERPNext yordamchingizman.\n"
                         "Savol bering, masalan:\n"
                         "• Kecha kim sales invoice kiritgan?\n"
                         "• Invoys qilinmagan delivery note'lar qancha?\n"
                         "• Bu oygi naqd kassa qancha?\n"
                         "• Balans / foyda-zarar (pnl) faylini ber\n"
                         "Suhbatni tozalash: /reset")
        return
    if cmd == "/reset":
        with SESSIONS_GUARD:
            SESSIONS.pop(chat_id, None)
            SESSION_TS.pop(chat_id, None)
        tg_send(chat_id, "Suhbat tozalandi.")
        return

    model = _pick_model(text)
    _log("SAVOL", f"user={user_id}", f"model={model}", repr(text[:200]))

    # Per-chat lock — shu chatning xabarlari ketma-ket ishlanadi
    with _get_chat_lock(chat_id):
        tg_send(chat_id, "🔎 Tekshiryapman...")
        with SESSIONS_GUARD:
            hist = SESSIONS.setdefault(chat_id, [])
            SESSION_TS[chat_id] = time.time()
        hist.append({"role": "user", "content": text})
        _trim_history(hist)
        try:
            answer = ask_claude(hist, chat_id, model=model)
        except Exception as e:
            # To'liq xatoni log'ga yozamiz, foydalanuvchiga sodda xabar beramiz
            _log("XATO", f"chat={chat_id}", repr(str(e)))
            traceback.print_exc()
            answer = ("Kechirasiz, javob tayyorlashda texnik xatolik yuz berdi. "
                      "Birozdan so'ng qayta urinib ko'ring.")
        with SESSIONS_GUARD:
            SESSION_TS[chat_id] = time.time()
    tg_send(chat_id, answer)


def main():
    print("Bot ishga tushdi. To'xtatish: Ctrl+C")
    _autodetect_company()

    # Eski (bot o'chiq turgan paytdagi) xabarlarni tashlab yuboramiz —
    # qayta ishga tushganda eski savollarga javob bermasligi uchun.
    try:
        r0 = requests.get(f"{ERP_TG_API}/getUpdates",
                          params={"timeout": 0, "offset": -1}, timeout=20)
        res0 = r0.json().get("result", [])
        offset = (res0[-1]["update_id"] + 1) if res0 else None
    except Exception:
        offset = None

    seen = set()   # qayta ishlangan update_id'lar (ikki marta javobning oldini oladi)

    while True:
        try:
            resp = requests.get(f"{ERP_TG_API}/getUpdates",
                                params={"offset": offset, "timeout": 30}, timeout=40)
            updates = resp.json().get("result", [])
            _prune_sessions()
            for upd in updates:
                uid = upd["update_id"]
                # offset'ni DARHOL oldinga suramiz — Telegram bu update'ni qayta yubormaydi
                offset = uid + 1
                if uid in seen:
                    continue
                seen.add(uid)
                if len(seen) > 1000:
                    seen.clear()
                    seen.add(uid)

                msg = upd.get("message") or {}
                if not msg.get("text"):
                    continue
                # Har xabarni alohida thread'da — uzun hisobot boshqa userlarni bloklamaydi
                threading.Thread(target=handle_message, args=(msg,), daemon=True).start()

        except KeyboardInterrupt:
            print("To'xtatildi.")
            break
        except Exception as e:
            print("Tsikl xatosi:", e)
            time.sleep(5)


if __name__ == "__main__":
    main()